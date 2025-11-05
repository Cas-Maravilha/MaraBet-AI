"""
Gerenciador de Exportação - MaraBet AI
Sistema completo de exportação de relatórios (PDF, CSV, Excel)
"""

import pandas as pd
import json
import os
import logging
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
from datetime import datetime
import base64
import io
from pathlib import Path

# Dependências opcionais para PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Dependências opcionais para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

@dataclass
class ExportConfig:
    """Configuração de exportação"""
    format: str
    filename: str
    include_charts: bool = True
    include_metadata: bool = True
    page_orientation: str = "portrait"
    font_size: int = 10
    margin: float = 1.0

@dataclass
class ExportResult:
    """Resultado da exportação"""
    success: bool
    filename: str
    file_path: str
    file_size: int
    format: str
    error_message: Optional[str] = None

class ExportManager:
    """
    Gerenciador de exportação de relatórios para MaraBet AI
    Suporta PDF, CSV, Excel e outros formatos
    """
    
    def __init__(self, output_dir: str = "accessibility/exports"):
        """
        Inicializa o gerenciador de exportação
        
        Args:
            output_dir: Diretório para arquivos exportados
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        # Verificar dependências
        self.pdf_available = REPORTLAB_AVAILABLE
        self.excel_available = OPENPYXL_AVAILABLE
        
        if not self.pdf_available:
            logger.warning("⚠️ ReportLab não disponível - exportação PDF desabilitada")
        if not self.excel_available:
            logger.warning("⚠️ OpenPyXL não disponível - exportação Excel desabilitada")
        
        logger.info("ExportManager inicializado")
    
    def export_to_pdf(self, 
                     data: Union[Dict, List, pd.DataFrame],
                     config: ExportConfig) -> ExportResult:
        """
        Exporta dados para PDF
        
        Args:
            data: Dados para exportar
            config: Configuração de exportação
            
        Returns:
            Resultado da exportação
        """
        try:
            if not self.pdf_available:
                return ExportResult(
                    success=False,
                    filename=config.filename,
                    file_path="",
                    file_size=0,
                    format="PDF",
                    error_message="ReportLab não disponível"
                )
            
            file_path = os.path.join(self.output_dir, f"{config.filename}.pdf")
            
            # Configurar documento
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4 if config.page_orientation == "portrait" else letter,
                rightMargin=config.margin * inch,
                leftMargin=config.margin * inch,
                topMargin=config.margin * inch,
                bottomMargin=config.margin * inch
            )
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Conteúdo
            story = []
            
            # Título
            story.append(Paragraph("Relatório MaraBet AI", title_style))
            story.append(Spacer(1, 20))
            
            # Metadados
            if config.include_metadata:
                metadata = self._generate_metadata()
                story.append(Paragraph("Informações do Relatório", styles['Heading2']))
                for key, value in metadata.items():
                    story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Dados
            if isinstance(data, pd.DataFrame):
                story.extend(self._dataframe_to_pdf_elements(data, styles))
            elif isinstance(data, dict):
                story.extend(self._dict_to_pdf_elements(data, styles))
            elif isinstance(data, list):
                story.extend(self._list_to_pdf_elements(data, styles))
            
            # Construir PDF
            doc.build(story)
            
            file_size = os.path.getsize(file_path)
            
            logger.info(f"✅ PDF exportado: {file_path}")
            return ExportResult(
                success=True,
                filename=f"{config.filename}.pdf",
                file_path=file_path,
                file_size=file_size,
                format="PDF"
            )
            
        except Exception as e:
            logger.error(f"❌ Erro ao exportar PDF: {e}")
            return ExportResult(
                success=False,
                filename=config.filename,
                file_path="",
                file_size=0,
                format="PDF",
                error_message=str(e)
            )
    
    def export_to_csv(self, 
                     data: Union[Dict, List, pd.DataFrame],
                     config: ExportConfig) -> ExportResult:
        """
        Exporta dados para CSV
        
        Args:
            data: Dados para exportar
            config: Configuração de exportação
            
        Returns:
            Resultado da exportação
        """
        try:
            file_path = os.path.join(self.output_dir, f"{config.filename}.csv")
            
            # Converter dados para DataFrame se necessário
            if isinstance(data, dict):
                df = pd.DataFrame([data])
            elif isinstance(data, list):
                if data and isinstance(data[0], dict):
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame(data, columns=['Value'])
            elif isinstance(data, pd.DataFrame):
                df = data
            else:
                df = pd.DataFrame([{"Value": str(data)}])
            
            # Adicionar metadados se solicitado
            if config.include_metadata:
                metadata = self._generate_metadata()
                for key, value in metadata.items():
                    df[f"Metadata_{key}"] = value
            
            # Exportar para CSV
            df.to_csv(file_path, index=False, encoding='utf-8')
            
            file_size = os.path.getsize(file_path)
            
            logger.info(f"✅ CSV exportado: {file_path}")
            return ExportResult(
                success=True,
                filename=f"{config.filename}.csv",
                file_path=file_path,
                file_size=file_size,
                format="CSV"
            )
            
        except Exception as e:
            logger.error(f"❌ Erro ao exportar CSV: {e}")
            return ExportResult(
                success=False,
                filename=config.filename,
                file_path="",
                file_size=0,
                format="CSV",
                error_message=str(e)
            )
    
    def export_to_excel(self, 
                       data: Union[Dict, List, pd.DataFrame],
                       config: ExportConfig) -> ExportResult:
        """
        Exporta dados para Excel
        
        Args:
            data: Dados para exportar
            config: Configuração de exportação
            
        Returns:
            Resultado da exportação
        """
        try:
            if not self.excel_available:
                return ExportResult(
                    success=False,
                    filename=config.filename,
                    file_path="",
                    file_size=0,
                    format="Excel",
                    error_message="OpenPyXL não disponível"
                )
            
            file_path = os.path.join(self.output_dir, f"{config.filename}.xlsx")
            
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório MaraBet AI"
            
            # Adicionar metadados se solicitado
            if config.include_metadata:
                metadata = self._generate_metadata()
                row = 1
                for key, value in metadata.items():
                    ws[f"A{row}"] = key
                    ws[f"B{row}"] = value
                    row += 1
                row += 1
            
            # Adicionar dados
            if isinstance(data, pd.DataFrame):
                self._dataframe_to_excel(data, ws, row)
            elif isinstance(data, dict):
                self._dict_to_excel(data, ws, row)
            elif isinstance(data, list):
                self._list_to_excel(data, ws, row)
            
            # Aplicar formatação
            self._apply_excel_formatting(ws)
            
            # Salvar arquivo
            wb.save(file_path)
            
            file_size = os.path.getsize(file_path)
            
            logger.info(f"✅ Excel exportado: {file_path}")
            return ExportResult(
                success=True,
                filename=f"{config.filename}.xlsx",
                file_path=file_path,
                file_size=file_size,
                format="Excel"
            )
            
        except Exception as e:
            logger.error(f"❌ Erro ao exportar Excel: {e}")
            return ExportResult(
                success=False,
                filename=config.filename,
                file_path="",
                file_size=0,
                format="Excel",
                error_message=str(e)
            )
    
    def export_to_json(self, 
                      data: Union[Dict, List, pd.DataFrame],
                      config: ExportConfig) -> ExportResult:
        """
        Exporta dados para JSON
        
        Args:
            data: Dados para exportar
            config: Configuração de exportação
            
        Returns:
            Resultado da exportação
        """
        try:
            file_path = os.path.join(self.output_dir, f"{config.filename}.json")
            
            # Preparar dados
            if isinstance(data, pd.DataFrame):
                export_data = data.to_dict('records')
            else:
                export_data = data
            
            # Adicionar metadados se solicitado
            if config.include_metadata:
                metadata = self._generate_metadata()
                export_data = {
                    "metadata": metadata,
                    "data": export_data
                }
            
            # Exportar para JSON
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            file_size = os.path.getsize(file_path)
            
            logger.info(f"✅ JSON exportado: {file_path}")
            return ExportResult(
                success=True,
                filename=f"{config.filename}.json",
                file_path=file_path,
                file_size=file_size,
                format="JSON"
            )
            
        except Exception as e:
            logger.error(f"❌ Erro ao exportar JSON: {e}")
            return ExportResult(
                success=False,
                filename=config.filename,
                file_path="",
                file_size=0,
                format="JSON",
                error_message=str(e)
            )
    
    def export_multiple_formats(self, 
                               data: Union[Dict, List, pd.DataFrame],
                               base_filename: str,
                               formats: List[str] = None) -> Dict[str, ExportResult]:
        """
        Exporta dados para múltiplos formatos
        
        Args:
            data: Dados para exportar
            base_filename: Nome base do arquivo
            formats: Lista de formatos (padrão: ['csv', 'json', 'pdf', 'excel'])
            
        Returns:
            Dicionário com resultados por formato
        """
        try:
            if formats is None:
                formats = ['csv', 'json']
                if self.pdf_available:
                    formats.append('pdf')
                if self.excel_available:
                    formats.append('excel')
            
            results = {}
            
            for format_type in formats:
                config = ExportConfig(
                    format=format_type,
                    filename=base_filename
                )
                
                if format_type == 'csv':
                    result = self.export_to_csv(data, config)
                elif format_type == 'json':
                    result = self.export_to_json(data, config)
                elif format_type == 'pdf':
                    result = self.export_to_pdf(data, config)
                elif format_type == 'excel':
                    result = self.export_to_excel(data, config)
                else:
                    result = ExportResult(
                        success=False,
                        filename=base_filename,
                        file_path="",
                        file_size=0,
                        format=format_type,
                        error_message=f"Formato não suportado: {format_type}"
                    )
                
                results[format_type] = result
            
            return results
            
        except Exception as e:
            logger.error(f"❌ Erro na exportação múltipla: {e}")
            return {}
    
    def _generate_metadata(self) -> Dict[str, str]:
        """Gera metadados para exportação"""
        return {
            "Gerado em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "Sistema": "MaraBet AI",
            "Versão": "1.0",
            "Formato": "Relatório de Análise"
        }
    
    def _dataframe_to_pdf_elements(self, df: pd.DataFrame, styles) -> List:
        """Converte DataFrame para elementos PDF"""
        elements = []
        
        # Título da tabela
        elements.append(Paragraph("Dados do Relatório", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        # Converter DataFrame para tabela
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _dict_to_pdf_elements(self, data: Dict, styles) -> List:
        """Converte dicionário para elementos PDF"""
        elements = []
        
        elements.append(Paragraph("Informações do Relatório", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        for key, value in data.items():
            elements.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
            elements.append(Spacer(1, 6))
        
        return elements
    
    def _list_to_pdf_elements(self, data: List, styles) -> List:
        """Converte lista para elementos PDF"""
        elements = []
        
        elements.append(Paragraph("Lista de Dados", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        for i, item in enumerate(data, 1):
            elements.append(Paragraph(f"{i}. {item}", styles['Normal']))
            elements.append(Spacer(1, 6))
        
        return elements
    
    def _dataframe_to_excel(self, df: pd.DataFrame, ws, start_row: int):
        """Converte DataFrame para Excel"""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
    def _dict_to_excel(self, data: Dict, ws, start_row: int):
        """Converte dicionário para Excel"""
        for key, value in data.items():
            ws.append([key, value])
    
    def _list_to_excel(self, data: List, ws, start_row: int):
        """Converte lista para Excel"""
        for i, item in enumerate(data, 1):
            ws.append([i, item])
    
    def _apply_excel_formatting(self, ws):
        """Aplica formatação ao worksheet Excel"""
        # Formatação do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_export_ui(self) -> str:
        """Cria interface de usuário para exportação"""
        return """
<div class="export-manager" role="region" aria-label="Gerenciador de Exportação">
    <h3>Exportar Relatório</h3>
    
    <div class="export-form">
        <div class="form-group">
            <label for="export-filename">Nome do arquivo:</label>
            <input type="text" id="export-filename" class="form-control" placeholder="relatorio_marabet" value="relatorio_marabet">
        </div>
        
        <div class="form-group">
            <label>Formatos disponíveis:</label>
            <div class="format-options">
                <label class="format-option">
                    <input type="checkbox" id="format-csv" checked>
                    <span>CSV</span>
                </label>
                <label class="format-option">
                    <input type="checkbox" id="format-json" checked>
                    <span>JSON</span>
                </label>
                <label class="format-option">
                    <input type="checkbox" id="format-pdf">
                    <span>PDF</span>
                </label>
                <label class="format-option">
                    <input type="checkbox" id="format-excel">
                    <span>Excel</span>
                </label>
            </div>
        </div>
        
        <div class="form-group">
            <label class="checkbox-option">
                <input type="checkbox" id="include-metadata" checked>
                <span>Incluir metadados</span>
            </label>
        </div>
        
        <div class="form-group">
            <button class="btn btn-primary" onclick="exportData()">
                <i class="fas fa-download"></i> Exportar
            </button>
            <button class="btn btn-secondary" onclick="previewExport()">
                <i class="fas fa-eye"></i> Visualizar
            </button>
        </div>
    </div>
    
    <div id="export-status" class="export-status" style="display: none;">
        <div class="status-content">
            <div class="spinner"></div>
            <span>Exportando...</span>
        </div>
    </div>
</div>

<style>
.export-manager {
    background: var(--surface-color);
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 20px;
    margin: 20px 0;
}

.export-form {
    display: grid;
    gap: 15px;
}

.form-group {
    display: flex;
    flex-direction: column;
    gap: 5px;
}

.form-group label {
    font-weight: 500;
    color: var(--text-color);
}

.form-control {
    padding: 8px 12px;
    border: 1px solid var(--border-color);
    border-radius: 4px;
    background: var(--background-color);
    color: var(--text-color);
}

.format-options {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(100px, 1fr));
    gap: 10px;
}

.format-option {
    display: flex;
    align-items: center;
    gap: 5px;
    cursor: pointer;
}

.checkbox-option {
    display: flex;
    align-items: center;
    gap: 5px;
    cursor: pointer;
}

.btn {
    padding: 10px 20px;
    border: none;
    border-radius: 4px;
    cursor: pointer;
    font-weight: 500;
    display: inline-flex;
    align-items: center;
    gap: 5px;
}

.btn-primary {
    background: var(--primary-color);
    color: white;
}

.btn-secondary {
    background: var(--text-secondary-color);
    color: white;
}

.export-status {
    margin-top: 20px;
    padding: 15px;
    background: var(--info-color)20;
    border: 1px solid var(--info-color);
    border-radius: 4px;
}

.status-content {
    display: flex;
    align-items: center;
    gap: 10px;
}

.spinner {
    width: 20px;
    height: 20px;
    border: 2px solid var(--border-color);
    border-top: 2px solid var(--primary-color);
    border-radius: 50%;
    animation: spin 1s linear infinite;
}

@keyframes spin {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}
</style>

<script>
function exportData() {
    const filename = document.getElementById('export-filename').value;
    const formats = [];
    
    if (document.getElementById('format-csv').checked) formats.push('csv');
    if (document.getElementById('format-json').checked) formats.push('json');
    if (document.getElementById('format-pdf').checked) formats.push('pdf');
    if (document.getElementById('format-excel').checked) formats.push('excel');
    
    const includeMetadata = document.getElementById('include-metadata').checked;
    
    if (formats.length === 0) {
        alert('Selecione pelo menos um formato');
        return;
    }
    
    // Mostrar status
    document.getElementById('export-status').style.display = 'block';
    
    // Simular exportação (substituir por chamada real à API)
    setTimeout(() => {
        document.getElementById('export-status').style.display = 'none';
        alert('Exportação concluída!');
    }, 2000);
}

function previewExport() {
    alert('Visualização de exportação (implementar)');
}
</script>
        """
    
    def get_export_history(self) -> List[Dict[str, Any]]:
        """Retorna histórico de exportações"""
        try:
            history = []
            
            for file in os.listdir(self.output_dir):
                file_path = os.path.join(self.output_dir, file)
                if os.path.isfile(file_path):
                    stat = os.stat(file_path)
                    history.append({
                        "filename": file,
                        "size": stat.st_size,
                        "created": datetime.fromtimestamp(stat.st_ctime),
                        "modified": datetime.fromtimestamp(stat.st_mtime)
                    })
            
            return sorted(history, key=lambda x: x["created"], reverse=True)
            
        except Exception as e:
            logger.error(f"❌ Erro ao obter histórico: {e}")
            return []
    
    def cleanup_old_exports(self, days: int = 30):
        """Remove exportações antigas"""
        try:
            cutoff_date = datetime.now().timestamp() - (days * 24 * 60 * 60)
            
            for file in os.listdir(self.output_dir):
                file_path = os.path.join(self.output_dir, file)
                if os.path.isfile(file_path):
                    if os.path.getctime(file_path) < cutoff_date:
                        os.remove(file_path)
                        logger.info(f"✅ Arquivo antigo removido: {file}")
            
        except Exception as e:
            logger.error(f"❌ Erro na limpeza: {e}")
